#!/usr/bin/env python
import os
import sys
import typing
import warnings
from datetime import date

import yaml
from mysql.connector import connection
from openpyxl import load_workbook

import json


def main():
    with open("../config.yml", "r") as yamlfile:
        config = yaml.load(yamlfile, Loader=yaml.FullLoader)
        print("Read successful")

    # today = str(date.today())
    cnx = connection.MySQLConnection(
        user=config["username"],
        password=config["password"],
        host="anmeldung.worldscoutjamboree.de",
        port=config["port"],
        database=config["database"],
    )

    cursor = cnx.cursor()
    
    # load excel file
    workbook = load_workbook(filename="2023-04-20--Download-Korea.xlsx")
    worksheet = workbook.active
    k_id_column = 3
    url_column = 27
 
    for row in worksheet.iter_rows():
        korea_id = row[k_id_column].value
        id = str(row[url_column].value).split("/")[-1].replace(".html","")
        update = f"update people set korea_id='{korea_id}' where id='{id}'"
        print(update)
        cursor.execute(update)
        cnx.commit()
        print (cursor.rowcount, "record(s) affected")

if __name__ == "__main__":
    sys.exit(main())
